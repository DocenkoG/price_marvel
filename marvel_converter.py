# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import io
import sys
import configparser
import time
import openpyxl                       # Для .xlsx
# import xlrd                         # для .xls
from   price_tools import getCellXlsx, quoted, dump_cell, currencyType, subInParentheses



def convert2csv( myname ):
    global log
    global SheetName
    global FilenameIn
    global FilenameOut
    global out_columns_names
    global out_columns_j
    global in_columns_j
    global colGrp
    global colSGrp
    global GrpFonti
    global BrandFonti
    global SubGrpFonti
    global HeaderFonti
    global HeaderFontSize
    global RegularFontSize
    global SubGrpBackgroundColor
    global GrpBackgroundColor
    global strHeader
    global SubGrpFontSize
    global GrpFontSize
    make_loger()
    log.debug('Begin ' + __name__ + ' convert2csv')

    # Прочитать конфигурацию из файла
    ff = config_read( myname )
    log.debug('Открываю файл '+ FilenameIn)
    book = openpyxl.load_workbook(filename = FilenameIn, read_only=False, keep_vba=False, data_only=True)
#   book = xlrd.open_workbook( FilenameIn.encode('cp1251'), formatting_info=True)
#   book = xlrd.open_workbook( os.path.join( mydir, FilenameIn.encode('cp1251')), formatting_info=True)
    
    log.debug('Устанавливаю страницу ' + SheetName )
#   sh = book.sheet_by_name( SheetName )                     # xls
    sh = book[SheetName]                                     # xlsx   
       
 
    ssss = []
    line_qty = 0
    log.debug('На странице %d строк' % book[SheetName].max_row)
                                                             # цикл по строкам файла
    for i in range(book[SheetName].min_row, book[SheetName].max_row+1) :
        i_last = i
        try:
            ccc = float(getCellXlsx(row=i, col=in_columns_j['цена'], isDigit='Y', sheet=sh)) 
            if ccc <= 0 :
                continue
            else :                                                          # Информационная строка
                sss = []                                                    # формируемая строка для вывода в файл
                for outColName in out_columns_names :
                    if outColName in out_columns_j :
                        if outColName in ('закупка','продажа','цена') :
                            ss = getCellXlsx(row=i, col=out_columns_j[outColName], isDigit='Y', sheet=sh) 
                        else:
                            ss = getCellXlsx(row=i, col=out_columns_j[outColName], isDigit='N', sheet=sh)
                    else : 
                        # вычисляемое поле
                        if   outColName == 'подгруппа' :
                            s2 = getCellXlsx(row=i, col=in_columns_j['категория2'], isDigit='N', sheet=sh) 
                            s3 = getCellXlsx(row=i, col=in_columns_j['категория3'], isDigit='N', sheet=sh) 
                            s4 = getCellXlsx(row=i, col=in_columns_j['категория4'], isDigit='N', sheet=sh) 
                            ss = s2
                            if s3 != '' :  ss = ss + ' / ' + s3
                            if s4 != '' :  ss = ss + ' / ' + s4
                        elif outColName == 'наименование' :
                            s1 = getCellXlsx(row=i, col=in_columns_j['вендор'],       isDigit='N', sheet=sh) 
                            s2 = getCellXlsx(row=i, col=in_columns_j['наименование'], isDigit='N', sheet=sh) 
                            ss = s1 + ' ' + s2
                        else :
                            log.debug('Не определено вычисляемое поле: <' + outColName + '>' )
                    sss.append( quoted( ss))
                ssss.append(','.join(sss))
        except Exception as e:
            log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'<' + '>' )
            raise e

    
    f2 = open( FilenameOut, 'w', encoding='cp1251')
    f2.write(strHeader  + ',\n')
    data = ',\n'.join(ssss) +','
    dddd = data.encode(encoding='cp1251', errors='replace')
    data = dddd.decode(encoding='cp1251')
    f2.write(data)
    f2.close()



def config_read( myname ):
    global log
    global SheetName
    global FilenameIn
    global FilenameOut
    global out_columns_names
    global out_columns_j
    global in_columns_j
    global colGrp
    global colSGrp
    global GrpFonti
    global SubGrpFonti
    global BrandFonti
    global HeaderFonti
    global HeaderFontSize
    global RegularFontSize
    global SubGrpBackgroundColor
    global GrpBackgroundColor
    global strHeader
    global SubGrpFontSize
    global GrpFontSize

    cfgFName = myname + '.cfg'
    log.debug('Begin config_read ' + cfgFName )
    
    config = configparser.ConfigParser()
    if os.path.exists(cfgFName):     config.read( cfgFName)
    else : log.debug('Не найден файл конфигурации.')

    # в разделе [cols_in] находится список интересующих нас колонок и номера столбцов исходного файла
    in_columns_names = config.options('cols_in')
    in_columns_j = {}
    for vName in in_columns_names :
        if ('' != config.get('cols_in', vName)) :
            in_columns_j[vName] = config.getint('cols_in', vName) 
    
    # По разделу [cols_out] формируем перечень выводимых колонок и строку заголовка результирующего CSV файла
    temp_list = config.options('cols_out')
    temp_list.sort()

    out_columns_names = []
    for vName in temp_list :
        if ('' != config.get('cols_out', vName)) :
            out_columns_names.append(vName)
    
    out_columns_j = {}
    for vName in out_columns_names :
        tName = config.get('cols_out', vName)
        if  tName in in_columns_j :
            out_columns_j[vName] = in_columns_j[tName]
    print('-----------------------------------')
    for vName in out_columns_j :
        print(vName, '\t', out_columns_j[vName])    
    print('-----------------------------------')
    strHeader = ','.join(out_columns_names)           # +',бренд,группа,подгруппа'
    print('HEAD =', strHeader)

    # считываем имена файлов и имя листа
    FilenameIn   = config.get('input','Filename_in' )
    SheetName    = config.get('input','SheetName'   )      
    FilenameOut  = config.get('input','Filename_out')
    print('SHEET=', SheetName)
    
    # считываем признаки группы и подгруппы
    if ('' != config.get('grp_properties',  'группа')) :
        colGrp               = config.getint('grp_properties',     'группа')
    if ('' != config.get('grp_properties',  'подгруппа')) :
        colSGrp              = config.getint('grp_properties',  'подгруппа')
    if ('' != config.get('grp_properties',  'GrpFonti')) :
        GrpFonti             = config.getint('grp_properties',   'GrpFonti')
    if ('' != config.get('grp_properties',  'SubGrpFonti')) :
        SubGrpFonti          = config.getint('grp_properties','SubGrpFonti')
    if ('' != config.get('grp_properties',  'BrandFonti')) :
        BrandFonti           = config.getint('grp_properties', 'BrandFonti')
    if ('' != config.get('grp_properties',  'HeaderFonti')) :
        HeaderFonti          = config.getint('grp_properties','HeaderFonti')
    if ('' != config.get('grp_properties',  'HeaderFontSize')) :
        HeaderFontSize       = config.getint('grp_properties','HeaderFontSize')
    if ('' != config.get('grp_properties',  'RegularFontSize')) :
        RegularFontSize      = config.getint('grp_properties','RegularFontSize')
    if ('' != config.get('grp_properties',  'SubGrpFontSize')): 
        SubGrpFontSize       = config.getint('grp_properties','SubGrpFontSize')
    if ('' != config.get('grp_properties',  'GrpFontSize')) :
        GrpFontSize          = config.getint('grp_properties',   'GrpFontSize')
    if ('' != config.get('grp_properties',  'SubGrpBackgroundColor')) :
        SubGrpBackgroundColor= config.getint('grp_properties','SubGrpBackgroundColor')
    if ('' != config.get('grp_properties',  'GrpBackgroundColor')) :
        GrpBackgroundColor   = config.getint('grp_properties',   'GrpBackgroundColor')
    subgrpfontbold           = config.get('grp_properties','subgrpfontbold')
    grpfontbold              = config.get('grp_properties',   'grpfontbold')
    return FilenameIn




def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')